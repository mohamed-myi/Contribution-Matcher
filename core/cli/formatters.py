# Output formatters for various formats

import csv
import json

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import markdown  # noqa: F401

    HAS_MARKDOWN = True
except ImportError:
    HAS_MARKDOWN = False


def format_text(issues: list[dict], verbose: bool = False) -> str:
    """
    Format issues as plain text.

    Returns - Formatted text string
    """
    if not issues:
        return "No issues found.\n"

    output = []
    for i, issue in enumerate(issues, 1):
        output.append(f"\n{i}. {issue.get('title', 'N/A')}")
        output.append(f"   URL: {issue.get('url', 'N/A')}")
        if verbose:
            output.append(
                f"   Repo: {issue.get('repo_owner', 'N/A')}/{issue.get('repo_name', 'N/A')}"
            )
            output.append(f"   Type: {issue.get('issue_type', 'N/A')}")
            output.append(f"   Difficulty: {issue.get('difficulty', 'N/A')}")
            if issue.get("score"):
                output.append(f"   Score: {issue.get('score', 'N/A')}")
        output.append("")

    return "\n".join(output)


def format_json(issues: list[dict]) -> str:
    """
    Format issues as JSON.

    Returns - JSON string
    """
    return json.dumps(issues, indent=2, default=str)


def format_table(issues: list[dict], verbose: bool = False) -> str:
    """
    Format issues as a table.

    Returns - Table string
    """
    if not issues:
        return "No issues found.\n"

    # Determine columns
    if verbose:
        columns = ["#", "Title", "Repo", "Type", "Difficulty", "Score", "URL"]
    else:
        columns = ["#", "Title", "Score", "URL"]

    # Calculate column widths
    widths = {col: len(col) for col in columns}
    for issue in issues:
        title = str(issue.get("title", ""))[:50]
        widths["Title"] = max(widths.get("Title", 0), len(title))
        if verbose:
            repo = f"{issue.get('repo_owner', '')}/{issue.get('repo_name', '')}"
            widths["Repo"] = max(widths.get("Repo", 0), len(repo))
            widths["Type"] = max(widths.get("Type", 0), len(str(issue.get("issue_type", ""))))
            widths["Difficulty"] = max(
                widths.get("Difficulty", 0), len(str(issue.get("difficulty", "")))
            )
        if issue.get("score"):
            widths["Score"] = max(widths.get("Score", 0), len(str(issue.get("score", ""))))
        widths["URL"] = max(widths.get("URL", 0), 50)

    # Build table
    output = []

    # Header
    header = " | ".join(col.ljust(widths.get(col, len(col))) for col in columns)
    output.append(header)
    output.append("-" * len(header))

    # Rows
    for i, issue in enumerate(issues, 1):
        row = []
        row.append(str(i).ljust(widths.get("#", 3)))
        title = str(issue.get("title", ""))[:50]
        row.append(title.ljust(widths.get("Title", len(title))))
        if verbose:
            repo = f"{issue.get('repo_owner', '')}/{issue.get('repo_name', '')}"
            row.append(repo.ljust(widths.get("Repo", len(repo))))
            row.append(str(issue.get("issue_type", "")).ljust(widths.get("Type", 10)))
            row.append(str(issue.get("difficulty", "")).ljust(widths.get("Difficulty", 12)))
        if issue.get("score"):
            row.append(str(issue.get("score", "")).ljust(widths.get("Score", 6)))
        url = str(issue.get("url", ""))[:50]
        row.append(url.ljust(widths.get("URL", len(url))))
        output.append(" | ".join(row))

    return "\n".join(output) + "\n"


def format_csv(issues: list[dict], output_file: str | None = None) -> str | None:
    """
    Format issues as CSV.

    Returns - CSV string if output_file is None, otherwise writes to file
    """
    if not issues:
        return "" if output_file is None else None

    fieldnames = [
        "title",
        "url",
        "repo_owner",
        "repo_name",
        "issue_type",
        "difficulty",
        "time_estimate",
        "score",
    ]

    if output_file:
        with open(output_file, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
            writer.writeheader()
            for issue in issues:
                writer.writerow(issue)
        return None
    else:
        import io

        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for issue in issues:
            writer.writerow(issue)
        return output.getvalue()


def format_excel(issues: list[dict], output_file: str) -> None:
    """
    Format issues as Excel file.

    Raises:
        ImportError if openpyxl is not installed
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel export. Install with: pip install openpyxl"
        )

    if not issues:
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Issues"

    # Headers
    headers = [
        "Title",
        "URL",
        "Repo Owner",
        "Repo Name",
        "Issue Type",
        "Difficulty",
        "Time Estimate",
        "Score",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Data
    for row_idx, issue in enumerate(issues, 2):
        ws.cell(row=row_idx, column=1, value=issue.get("title", ""))
        ws.cell(row=row_idx, column=2, value=issue.get("url", ""))
        ws.cell(row=row_idx, column=3, value=issue.get("repo_owner", ""))
        ws.cell(row=row_idx, column=4, value=issue.get("repo_name", ""))
        ws.cell(row=row_idx, column=5, value=issue.get("issue_type", ""))
        ws.cell(row=row_idx, column=6, value=issue.get("difficulty", ""))
        ws.cell(row=row_idx, column=7, value=issue.get("time_estimate", ""))
        ws.cell(row=row_idx, column=8, value=issue.get("score", ""))

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in ws[column_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_file)


def format_markdown(issues: list[dict]) -> str:
    """
    Format issues as Markdown.

    Returns - Markdown string
    """
    if not issues:
        return "No issues found.\n"

    output = []
    output.append("| # | Title | Repo | Type | Difficulty | Score | URL |")
    output.append("|" + "|".join(["---"] * 7) + "|")

    for i, issue in enumerate(issues, 1):
        title = issue.get("title", "N/A").replace("|", "\\|")
        repo = f"{issue.get('repo_owner', 'N/A')}/{issue.get('repo_name', 'N/A')}"
        issue_type = issue.get("issue_type", "N/A")
        difficulty = issue.get("difficulty", "N/A")
        score = issue.get("score", "N/A")
        url = issue.get("url", "N/A")
        output.append(f"| {i} | {title} | {repo} | {issue_type} | {difficulty} | {score} | {url} |")

    return "\n".join(output) + "\n"


def format_html(issues: list[dict]) -> str:
    """
    Format issues as HTML table.

    Returns - HTML string
    """
    if not issues:
        return "<p>No issues found.</p>\n"

    output = []
    output.append("<table>")
    output.append("<thead>")
    output.append(
        "<tr><th>#</th><th>Title</th><th>Repo</th><th>Type</th><th>Difficulty</th><th>Score</th><th>URL</th></tr>"
    )
    output.append("</thead>")
    output.append("<tbody>")

    for i, issue in enumerate(issues, 1):
        title = issue.get("title", "N/A").replace("<", "&lt;").replace(">", "&gt;")
        repo = f"{issue.get('repo_owner', 'N/A')}/{issue.get('repo_name', 'N/A')}"
        issue_type = issue.get("issue_type", "N/A")
        difficulty = issue.get("difficulty", "N/A")
        score = issue.get("score", "N/A")
        url = issue.get("url", "N/A")
        output.append("<tr>")
        output.append(f"  <td>{i}</td>")
        output.append(f"  <td>{title}</td>")
        output.append(f"  <td>{repo}</td>")
        output.append(f"  <td>{issue_type}</td>")
        output.append(f"  <td>{difficulty}</td>")
        output.append(f"  <td>{score}</td>")
        output.append(f"  <td><a href='{url}'>{url[:50]}</a></td>")
        output.append("</tr>")

    output.append("</tbody>")
    output.append("</table>")

    return "\n".join(output) + "\n"


def format_output(
    issues: list[dict], format_type: str, verbose: bool = False, output_file: str | None = None
) -> str | None:
    """
    Format issues based on the specified format.

    Args:
        issues: List of issue dictionaries
        format_type: One of 'text', 'json', 'table', 'csv', 'excel', 'markdown', 'html'
        verbose: Whether to include detailed information
        output_file: Optional file path for file-based formats

    Returns:
        Formatted string (or None for file-based formats)
    """
    # Handle both string and mock objects
    if hasattr(format_type, "lower") or isinstance(format_type, str):
        format_type = format_type.lower()
    else:
        format_type = str(format_type).lower()

    if format_type == "text":
        return format_text(issues, verbose)
    elif format_type == "json":
        return format_json(issues)
    elif format_type == "table":
        return format_table(issues, verbose)
    elif format_type == "csv":
        if output_file:
            format_csv(issues, output_file)
            return None
        return format_csv(issues)
    elif format_type == "excel":
        if not output_file:
            raise ValueError("output_file is required for Excel format")
        format_excel(issues, output_file)
        return None
    elif format_type == "markdown":
        return format_markdown(issues)
    elif format_type == "html":
        return format_html(issues)
    else:
        raise ValueError(
            f"Unsupported format: {format_type}. Supported: text, json, table, csv, excel, markdown, html"
        )
